from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


# Paleta institucional alineada con las interfaces Flet/Streamlit.
GOB_GREEN = "1F4D3A"
GOB_GREEN_DARK = "163A2C"
GOB_GOLD = "B08D57"
GOB_CREAM = "F7F4EE"
TEXT_DARK = "202124"
BORDER_LIGHT = "E4E1DB"
WHITE = "FFFFFF"

FONT_NAME = "Montserrat"
FONT_SIZE = 11

HEADER_FILL = PatternFill(fill_type="solid", fgColor=WHITE)
ALT_ROW_FILL = PatternFill(fill_type="solid", fgColor=GOB_CREAM)

HEADER_FONT = Font(
    name=FONT_NAME,
    size=FONT_SIZE,
    bold=True,
    color=GOB_GREEN_DARK,
)
BODY_FONT = Font(
    name=FONT_NAME,
    size=FONT_SIZE,
    color=TEXT_DARK,
)

HEADER_ALIGNMENT = Alignment(
    horizontal="center",
    vertical="center",
    wrap_text=True,
)
BODY_ALIGNMENT = Alignment(
    vertical="center",
    wrap_text=False,
)

HEADER_BORDER = Border(
    bottom=Side(style="medium", color=GOB_GOLD),
)
BODY_BORDER = Border(
    bottom=Side(style="hair", color=BORDER_LIGHT),
)
