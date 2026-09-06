from __future__ import annotations

from typing import Any, Dict, List

from openpyxl.utils import get_column_letter

from exporters.excel.styles import (
    ALT_ROW_FILL,
    BODY_ALIGNMENT,
    BODY_BORDER,
    BODY_FONT,
    GOB_GREEN,
    HEADER_ALIGNMENT,
    HEADER_BORDER,
    HEADER_FILL,
    HEADER_FONT,
)


_MONEY_TOKENS = (
    "cargo",
    "abono",
    "saldo",
    "comision",
    "comisión",
    "interes",
    "interés",
    "isr",
)
_RATE_TOKENS = ("tasa", "gat")


def _number_format_for_header(header: str) -> str | None:
    normalized = header.strip().lower()
    if any(token in normalized for token in _RATE_TOKENS):
        return "0.00"
    if any(token in normalized for token in _MONEY_TOKENS):
        return '$#,##0.00;[Red]-$#,##0.00'
    return None


def write_table_sheet(
    ws,
    rows: List[Dict[str, Any]],
):
    """Escribe una colección de diccionarios como una hoja de Excel.

    Conserva exactamente las columnas y los valores normalizados del mapper;
    esta función sólo añade presentación visual, filtros y formatos de celda.
    """

    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 90
    ws.sheet_properties.tabColor = GOB_GREEN

    if not rows:
        ws.append(["Sin información"])
        cell = ws["A1"]
        cell.font = BODY_FONT
        cell.alignment = BODY_ALIGNMENT
        ws.column_dimensions["A"].width = 24
        return

    headers: list[str] = []
    for row in rows:
        for key in row:
            if key not in headers:
                headers.append(key)

    ws.append(headers)
    for row in rows:
        ws.append([row.get(header) for header in headers])

    ws.row_dimensions[1].height = 30
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = HEADER_BORDER

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for row_index, row in enumerate(ws.iter_rows(min_row=2), start=2):
        ws.row_dimensions[row_index].height = 21
        for cell in row:
            cell.font = BODY_FONT
            cell.alignment = BODY_ALIGNMENT
            cell.border = BODY_BORDER
            if row_index % 2 == 0:
                cell.fill = ALT_ROW_FILL

    for column_index, header in enumerate(headers, start=1):
        number_format = _number_format_for_header(header)
        if number_format:
            for cell in ws.iter_cols(
                min_col=column_index,
                max_col=column_index,
                min_row=2,
                max_row=ws.max_row,
            ):
                for item in cell:
                    if isinstance(item.value, (int, float)) and not isinstance(item.value, bool):
                        item.number_format = number_format

        values = [header]
        values.extend(
            cell.value
            for cell in next(
                ws.iter_cols(
                    min_col=column_index,
                    max_col=column_index,
                    min_row=2,
                    max_row=ws.max_row,
                )
            )
        )
        max_length = max(len(str(value)) if value is not None else 0 for value in values)
        width = min(max(max_length + 2, 11), 38)
        ws.column_dimensions[get_column_letter(column_index)].width = width
